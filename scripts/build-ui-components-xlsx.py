#!/usr/bin/env python3
"""Gera component-mapping/ui-components.xlsx a partir do JSON de mapeamento."""
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
DATA = json.loads((ROOT / "component-mapping" / "ui-components.json").read_text())
OUT = ROOT / "component-mapping" / "ui-components.xlsx"

FONT = "Arial"
HEADER_FILL = PatternFill("solid", start_color="1F2A44")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name=FONT, bold=True, size=14)
BASE_FONT = Font(name=FONT, size=10)
THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
NO_FILL = PatternFill("solid", start_color="FFF2CC")  # destaque "não usado"

ORIGIN_LABEL = {
    "local": "Local (devportal)",
    "external": "Externo (repo VTEX)",
    "third-party": "Terceiro (npm)",
}


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = BORDER


def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


wb = Workbook()

# --------------------------------------------------------------------------
# Aba 1: Resumo
# --------------------------------------------------------------------------
ws = wb.active
ws.title = "Resumo"
meta = DATA["meta"]
ws["A1"] = "Mapeamento de Componentes de UI — devportal"
ws["A1"].font = TITLE_FONT
ws["A3"] = "Gerado em:"
ws["B3"] = meta["generatedAt"]
ws["A4"] = "Arquivos varridos (src/):"
ws["B4"] = meta["filesScanned"]
ws["A5"] = "Total de componentes:"
ws["B5"] = meta["totalComponents"]
ws["A6"] = "Locais não usados (neste repo):"
ws["B6"] = meta["localUnusedCount"]
for r in range(3, 7):
    ws.cell(row=r, column=1).font = Font(name=FONT, bold=True, size=10)
    ws.cell(row=r, column=2).font = BASE_FONT

ws["A8"] = "Ressalva"
ws["A8"].font = Font(name=FONT, bold=True, size=11)
ws["A9"] = meta["caveat"]
ws["A9"].font = Font(name=FONT, size=10, italic=True)
ws["A9"].alignment = Alignment(wrap_text=True, vertical="top")
ws.merge_cells("A9:E11")

row = 13
ws.cell(row=row, column=1, value="Componentes por repositório").font = Font(name=FONT, bold=True, size=11)
row += 1
ws.cell(row=row, column=1, value="Repositório")
ws.cell(row=row, column=2, value="Qtd. componentes")
style_header(ws, row, 2)
row += 1
for repo, count in sorted(meta["countsByRepository"].items(), key=lambda x: -x[1]):
    ws.cell(row=row, column=1, value=repo).font = BASE_FONT
    ws.cell(row=row, column=2, value=count).font = BASE_FONT
    ws.cell(row=row, column=1).border = BORDER
    ws.cell(row=row, column=2).border = BORDER
    row += 1

autosize(ws, [42, 18, 14, 14, 14])

# --------------------------------------------------------------------------
# Aba 2: Componentes
# --------------------------------------------------------------------------
ws = wb.create_sheet("Componentes")
headers = [
    "Componente", "Tipo", "Kind", "Pacote", "Repositório",
    "Arquivo de declaração / origem", "Usado?", "Qtd. usos", "Usado em (arquivos)",
]
ws.append(headers)
style_header(ws, 1, len(headers))
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

for c in DATA["components"]:
    ws.append([
        c["name"],
        ORIGIN_LABEL.get(c["origin"], c["origin"]),
        c["kind"],
        c["package"] or "",
        c["repository"],
        c["declarationFile"] or c["importSource"] or "",
        "Sim" if c["used"] else "Não",
        c["usageCount"],
        "\n".join(c["usedInFiles"]),
    ])
    r = ws.max_row
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=r, column=col)
        cell.font = BASE_FONT
        cell.border = BORDER
        cell.alignment = Alignment(vertical="top", wrap_text=(col == 9))
    if not c["used"]:
        ws.cell(row=r, column=7).fill = NO_FILL

autosize(ws, [30, 20, 12, 22, 38, 50, 9, 10, 60])

# --------------------------------------------------------------------------
# Aba 3: Locais não usados
# --------------------------------------------------------------------------
ws = wb.create_sheet("Locais não usados")
ws["A1"] = "Componentes locais sem uso detectado em devportal/src"
ws["A1"].font = TITLE_FONT
ws["A2"] = "Atenção: podem ser consumidos pelo conteúdo MDX externo (vtexdocs/dev-portal-content), não verificado aqui."
ws["A2"].font = Font(name=FONT, size=10, italic=True)
ws["A2"].alignment = Alignment(wrap_text=True)
ws.merge_cells("A2:C2")
hdr = ["Componente", "Arquivo de declaração", "Observações"]
ws.append([])
ws.append(hdr)
style_header(ws, 4, len(hdr))
for c in DATA["components"]:
    if c["origin"] == "local" and not c["used"]:
        ws.append([c["name"], c["declarationFile"] or "", c["notes"] or ""])
        r = ws.max_row
        for col in range(1, len(hdr) + 1):
            ws.cell(row=r, column=col).font = BASE_FONT
            ws.cell(row=r, column=col).border = BORDER
autosize(ws, [30, 55, 55])

wb.save(OUT)
print(f"✓ XLSX: {OUT.relative_to(ROOT)}  ({len(DATA['components'])} componentes)")
